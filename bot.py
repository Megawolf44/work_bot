import logging
import sqlite3
import zipfile
import traceback
import shutil
import os
from datetime import datetime
from flask import Flask, send_from_directory
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InputFile
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters
)
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from dotenv import load_dotenv

# === Настройка окружения ===
load_dotenv()
TOKEN = os.getenv("TELEGRAM_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))
DOMAIN = os.getenv("DOMAIN", "http://127.0.0.1:4000")  # URL домена или локального сервера

EXCEL_FILE = 'requests.xlsx'
DB_FILE = 'requests.db'
FILES_DIR = "files"
os.makedirs(FILES_DIR, exist_ok=True)

logging.basicConfig(level=logging.INFO)

# === Состояния ===
WALL, SHROBE, AREA, PHOTOS, NAME, PHONE, ADDRESS, CONFIRM = range(8)
user_data = {}

# === SQLite ===
conn = sqlite3.connect(DB_FILE, check_same_thread=False)
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT, wall_type TEXT, shroblenie TEXT, area REAL,
    full_name TEXT, phone TEXT, address TEXT, total REAL, date TEXT)''')
conn.commit()

# === Excel ===
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата", "Username", "Тип стен", "Штробление", "Площадь",
               "ФИО", "Телефон", "Адрес", "Стоимость"])
    wb.save(EXCEL_FILE)

# === Telegram Handlers ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user_data[uid] = {}
    markup = ReplyKeyboardMarkup([
        ["Железобетон / Бетон"],
        ["Пазогребень / Пеноблок"],
        ["Каркас"]
    ], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("Выберите тип стен:", reply_markup=markup)
    return WALL

async def wall(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text
    user_data[uid]['wall'] = text
    if text == "Каркас":
        user_data[uid]['shrobe'] = "нет"
        await update.message.reply_text("Введите площадь помещения в м²:", reply_markup=ReplyKeyboardRemove())
        return AREA
    await update.message.reply_text("Нужно ли штробление? (Да/Нет)")
    return SHROBE

async def shrobe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text.lower()
    if text not in ["да", "нет"]:
        await update.message.reply_text("Пожалуйста, введите 'Да' или 'Нет'.")
        return SHROBE
    user_data[uid]['shrobe'] = text
    await update.message.reply_text("Введите площадь помещения в м²:")
    return AREA

async def area(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    try:
        area = float(update.message.text.replace(",", "."))
        user_data[uid]['area'] = area
        user_data[uid]['photos'] = []
        await update.message.reply_text("Отправьте до 5 фото объекта. После этого введите /done")
        return PHOTOS
    except ValueError:
        await update.message.reply_text("Введите корректное число.")
        return AREA

async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if len(user_data[uid]['photos']) >= 5:
        await update.message.reply_text("Максимум 5 фото.")
        return PHOTOS
    photo = update.message.photo[-1]
    file = await photo.get_file()
    path = f"{FILES_DIR}/{uid}_{len(user_data[uid]['photos'])}.jpg"
    await file.download_to_drive(path)
    user_data[uid]['photos'].append(path)

async def done_photos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите ваши Фамилию Имя Отчество:")
    return NAME

async def name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_data[update.effective_user.id]['full_name'] = update.message.text
    await update.message.reply_text("Введите номер телефона для связи:")
    return PHONE

async def phone_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_data[update.effective_user.id]['phone'] = update.message.text
    await update.message.reply_text("Введите адрес объекта для электромонтажных работ:")
    return ADDRESS

async def address_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user_data[uid]['address'] = update.message.text
    data = user_data[uid]
    wall, shrobe, area = data['wall'], data['shrobe'], data['area']
    price_per_m2 = 3500
    if wall == "Железобетон / Бетон":
        price_per_m2 = 4500 if shrobe == "да" else 3500
    elif wall == "Пазогребень / Пеноблок":
        price_per_m2 = 4000 if shrobe == "да" else 3500
    total = area * price_per_m2 + 5000
    data['total'] = total
    text = (
        f"Проверьте данные:\n\n"
        f"Тип стен: {wall}\nШтробление: {shrobe}\nПлощадь: {area} м²\n"
        f"ФИО: {data['full_name']}\nТелефон: {data['phone']}\nАдрес: {data['address']}\n"
        f"Стоимость: {total:.2f} ₽\n\nПодтвердите отправку?")
    markup = ReplyKeyboardMarkup([["Подтвердить", "Отменить"]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(text, reply_markup=markup)
    return CONFIRM

async def confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if update.message.text != "Подтвердить":
        await update.message.reply_text("Заявка отменена.", reply_markup=ReplyKeyboardRemove())
        user_data.pop(uid, None)
        return ConversationHandler.END

    data = user_data[uid]
    username = update.effective_user.username or "неизвестно"

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"), username, data['wall'],
            data['shrobe'], data['area'], data['full_name'], data['phone'],
            data['address'], data['total']
        ])
        wb.save(EXCEL_FILE)

        timestamp = int(datetime.now().timestamp())
        pdf_path = f"{FILES_DIR}/request_{uid}_{timestamp}.pdf"
        c = canvas.Canvas(pdf_path, pagesize=A4)
        c.setFont("Helvetica", 12)
        c.drawString(50, 800, f"ЗАЯВКА от {update.effective_user.first_name}")
        c.drawString(50, 780, f"ФИО: {data['full_name']}")
        c.drawString(50, 760, f"Телефон: {data['phone']}")
        c.drawString(50, 740, f"Адрес: {data['address']}")
        c.drawString(50, 720, f"Тип стен: {data['wall']}")
        c.drawString(50, 700, f"Штробление: {data['shrobe']}")
        c.drawString(50, 680, f"Площадь: {data['area']} м²")
        c.drawString(50, 660, f"Стоимость: {data['total']:.2f} ₽")
        c.showPage()
        c.save()

        zip_path = f"{FILES_DIR}/zayavka_{uid}_{timestamp}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            zipf.write(pdf_path, arcname=os.path.basename(pdf_path))
            for photo_path in data['photos']:
                zipf.write(photo_path, arcname=os.path.basename(photo_path))

        cursor.execute("""INSERT INTO requests 
            (username, wall_type, shroblenie, area, full_name, phone, address, total, date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (username, data['wall'], data['shrobe'], data['area'],
             data['full_name'], data['phone'], data['address'],
             data['total'], datetime.now().strftime("%Y-%m-%d %H:%M")))
        conn.commit()

        # Отправка ZIP
        await context.bot.send_document(
            chat_id=ADMIN_ID,
            document=InputFile(zip_path),
            caption=f"Новая заявка от @{username}\nСкачать: {DOMAIN}/files/{os.path.basename(zip_path)}")

        await update.message.reply_text("Заявка успешно отправлена!", reply_markup=ReplyKeyboardRemove())

    except Exception as e:
        logging.error("Ошибка при обработке заявки:\n" + traceback.format_exc())
        await update.message.reply_text("Произошла ошибка при отправке заявки.", reply_markup=ReplyKeyboardRemove())

    finally:
        for path in data.get('photos', []):
            if os.path.exists(path): os.remove(path)
        if os.path.exists(pdf_path): os.remove(pdf_path)
        user_data.pop(uid, None)
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_data.pop(update.effective_user.id, None)
    await update.message.reply_text("Диалог отменён.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# === Flask server to serve files ===
flask_app = Flask(__name__)

@flask_app.route('/files/<path:filename>')
def download_file(filename):
    return send_from_directory(FILES_DIR, filename, as_attachment=True)

# === Main ===
def main():
    import threading
    threading.Thread(target=flask_app.run, kwargs={'debug': False}).start()
    app = ApplicationBuilder().token(TOKEN).build()
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            WALL: [MessageHandler(filters.TEXT & ~filters.COMMAND, wall)],
            SHROBE: [MessageHandler(filters.TEXT & ~filters.COMMAND, shrobe)],
            AREA: [MessageHandler(filters.TEXT & ~filters.COMMAND, area)],
            PHOTOS: [
                MessageHandler(filters.PHOTO, photo_handler),
                CommandHandler("done", done_photos)
            ],
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, name_handler)],
            PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, phone_handler)],
            ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, address_handler)],
            CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    app.add_handler(conv_handler)
    app.run_polling()

if __name__ == "__main__":
    main()
